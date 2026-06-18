"""
SMN Climatological Normals Downloader - Plain Output
=====================================================
Same logic as descarga_estaciones_smn.py but outputs plain Excel files with
no styling (no colors, no column widths, no alignment). Use this when you
only need the raw data and formatting is not required.

Each workbook has one sheet per climate variable (e.g. Temperatura Media,
Precipitacion Total). Each sheet contains yearly data rows plus MINIMA,
MAXIMA, MEDIA, and DESV.ST statistics at the bottom.

Output filename: {station_id}_{status}_{last_year}.xlsx
    status      OPERANDO or SUSPENDIDA — extracted from the .txt file header.
    last_year   Last year with recorded data found in the file.

Configuration (edit at the top of the script):
    CSV_ENTRADA     Path to the CSV file with station IDs to download.
    COLUMNA_ID      Column name holding the station IDs.
                    Set to None to use the first column automatically.
    ESTADO_CLAVE    State code used in the SMN URL (e.g. 'mich' for Michoacan,
                    'jal' for Jalisco). Check the SMN site for other state codes.
    CARPETA_SALIDA  Folder where Excel files will be saved. Created automatically
                    if it does not exist.
    PAUSA_SEGUNDOS  Seconds to wait between requests. Increase if you get many
                    timeout errors.

Usage:
    1. Set CSV_ENTRADA, ESTADO_CLAVE, and CARPETA_SALIDA.
    2. python descarga_estaciones_smn_sin_formato.py

Requirements:
    pip install requests openpyxl
"""

import requests
import urllib3
import time
import os
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)  # SMN uses a self-signed SSL cert; this suppresses the warning printed on every request

# ------------------------------------------------------------------ CONFIGURACION

CSV_ENTRADA    = "estaciones.csv"
COLUMNA_ID     = None               # None = usa la primera columna
ESTADO_CLAVE   = "mich"
CARPETA_SALIDA = "estaciones_smn"
PAUSA_SEGUNDOS = 1

BASE_URL = (
    "https://smn.conagua.gob.mx/tools/RESOURCES/"
    "Normales_Climatologicas/Mensuales/{estado}/mes{id}.txt"
)  # {estado} and {id} are filled in at runtime via .format() inside descargar_txt

COLUMNAS = ["AÑO", "ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
            "JUL", "AGO", "SEP", "OCT", "NOV", "DIC", "ACUM", "PROM", "MESES"]

# The 4 summary statistics rows at the bottom of each variable block in the .txt file
# Must match exactly as they appear in the file (UTF-8 accented characters)
STAT_LABELS = {"MÍNIMA", "MÁXIMA", "MEDIA", "DESV.ST"}


# ------------------------------------------------------------------ HELPERS

def es_stat(texto):
    return texto.strip() in STAT_LABELS

def es_anio(texto):
    t = texto.strip()
    return t.isdigit() and len(t) == 4

def es_encabezado_columnas(texto):
    return texto.strip() == "AÑO"

def nombre_hoja_valido(nombre):
    for ch in r'\/?*[]:|':
        nombre = nombre.replace(ch, "")
    return nombre[:31].strip()

def es_nombre_variable(linea_raw, stripped):
    if not stripped:
        return False
    if "\t" in linea_raw:
        return False
    if ":" in stripped:
        return False
    if linea_raw.startswith(" "):
        return False
    if es_anio(stripped):
        return False
    if es_stat(stripped):
        return False
    if es_encabezado_columnas(stripped):
        return False
    return True


# ------------------------------------------------------------------ PARSER

def parsear_bloques(texto):
    """
    Devuelve dict: { nombre_variable: [ [col0..col15], ... ] }
    Cada lista incluye filas de datos + las 4 filas de estadísticas.
    Todos los valores son strings (igual que en el TXT original).
    """
    bloques       = {}
    nombre_actual = None
    filas_actual  = []

    def guardar():
        nonlocal nombre_actual, filas_actual  # nonlocal lets this inner function modify variables from the outer scope
        if nombre_actual is not None and filas_actual:
            bloques[nombre_actual] = filas_actual[:]
        nombre_actual = None
        filas_actual  = []

    for linea in texto.splitlines():
        linea_raw = linea.rstrip("\r\n")
        stripped  = linea_raw.strip()

        if not stripped:
            guardar()
            continue

        partes  = linea_raw.split("\t")
        primera = partes[0].strip()

        if es_encabezado_columnas(primera):
            continue

        if (es_anio(primera) or es_stat(primera)) and nombre_actual is not None:
            while len(partes) < len(COLUMNAS):
                partes.append("")
            partes = partes[:len(COLUMNAS)]
            filas_actual.append([p.strip() for p in partes])
            continue

        if es_nombre_variable(linea_raw, stripped):
            guardar()
            nombre_actual = stripped
            filas_actual  = []

    guardar()
    return bloques


# ------------------------------------------------------------------ METADATA

def extraer_situacion(texto):
    for linea in texto.splitlines():
        if "SITUACI" in linea.upper() and ":" in linea:
            valor = linea.split(":", 1)[1].strip()
            return valor.split()[0].upper() if valor.split() else "DESCONOCIDA"
    return "DESCONOCIDA"

def extraer_ultimo_anio(bloques):
    ultimo = 0
    for filas in bloques.values():
        for fila in filas:
            primera = fila[0].strip()
            if primera.isdigit() and len(primera) == 4:
                ultimo = max(ultimo, int(primera))
    return str(ultimo) if ultimo else "SANANIO"


# ------------------------------------------------------------------ EXCEL

def crear_excel(station_id, bloques, situacion, ultimo_anio):
    wb = Workbook()
    wb.remove(wb.active)

    for nombre_var, filas in bloques.items():
        ws = wb.create_sheet(title=nombre_hoja_valido(nombre_var))

        # Fila 1: nombre de la variable
        ws.cell(row=1, column=1, value=nombre_var)

        # Fila 2: encabezados de columna
        for c, col in enumerate(COLUMNAS, start=1):
            ws.cell(row=2, column=c, value=col)

        # Filas de datos
        for r, fila in enumerate(filas, start=3):
            for c, val in enumerate(fila, start=1):
                ws.cell(row=r, column=c, value=val)

    archivo = os.path.join(CARPETA_SALIDA, f"{station_id}_{situacion}_{ultimo_anio}.xlsx")
    wb.save(archivo)
    return archivo


# ------------------------------------------------------------------ DESCARGA

def descargar_txt(station_id, estado, session, reintentos=4, espera=5):
    url = BASE_URL.format(estado=estado, id=str(station_id).zfill(5))  # zfill pads the ID with leading zeros to 5 digits (e.g. '123' -> '00123')
    for intento in range(1, reintentos + 1):
        try:
            r = session.get(url, timeout=30, verify=False)  # verify=False skips SSL cert check (needed for SMN's self-signed cert)
            if r.status_code == 200 and len(r.content) > 0:
                return r.content.decode("utf-8")
            print(f"  X  {station_id}  ->  HTTP {r.status_code}")
            return None
        except requests.exceptions.Timeout:
            if intento < reintentos:
                print(f"  !  {station_id}  ->  Timeout, reintentando ({intento}/{reintentos - 1})...")
                time.sleep(espera * intento)   # incremental backoff: waits 5s, then 10s, then 15s — gives the server more time on repeated failures
            else:
                print(f"  X  {station_id}  ->  Timeout tras {reintentos} intentos.")
                return None
        except requests.RequestException as e:
            print(f"  X  {station_id}  ->  {e}")
            return None


# ------------------------------------------------------------------ MAIN

def main():
    if not os.path.exists(CSV_ENTRADA):
        print(f"No se encontro '{CSV_ENTRADA}'.")
        return

    with open(CSV_ENTRADA, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader)
        col_idx = 0 if COLUMNA_ID is None else headers.index(COLUMNA_ID)
        ids = list({row[col_idx].strip() for row in reader
                    if row and row[col_idx].strip()})  # set comprehension automatically removes duplicate station IDs
    ids.sort()
    print(f"\n{len(ids)} estaciones encontradas.\n")

    os.makedirs(CARPETA_SALIDA, exist_ok=True)
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0"})  # mimics a browser request; some servers block requests with no User-Agent

    ok = err = 0

    for sid in ids:
        print(f"  Descargando {sid}...")
        texto = descargar_txt(sid, ESTADO_CLAVE, session)

        if texto is None:
            err += 1
            time.sleep(PAUSA_SEGUNDOS)
            continue

        bloques = parsear_bloques(texto)

        if not bloques:
            print(f"  AVISO {sid}: no se encontraron datos.")
            err += 1
            time.sleep(PAUSA_SEGUNDOS)
            continue

        situacion   = extraer_situacion(texto)
        ultimo_anio = extraer_ultimo_anio(bloques)

        try:
            archivo = crear_excel(sid, bloques, situacion, ultimo_anio)
            print(f"  OK  {sid}  ->  {len(bloques)} hojas  ->  {archivo}")
            ok += 1
        except Exception as e:
            print(f"  ERROR  {sid}  ->  {e}")
            err += 1

        time.sleep(PAUSA_SEGUNDOS)

    print(f"\nListo. {ok} estaciones guardadas, {err} con error.")
    print(f"Archivos en: ./{CARPETA_SALIDA}/")


if __name__ == "__main__":
    main()
